from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import sys

def crear_grafica_desde_excel(archivo_excel):
    # Cargar el libro de trabajo Excel
    wb = load_workbook(archivo_excel)
    
    # Obtener la hoja "Grafica"
    ws = wb['Grafica']
    
    # Obtener la dimensión real de los datos
    dimension = ws.calculate_dimension()
    if dimension:
        # Obtener los datos y las categorías
        data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=ws.max_row)
        
        # Crear un objeto BarChart
        chart = BarChart()
        chart.title = "Gráfica"
        chart.x_axis.title = ws.cell(row=1, column=1).value
        chart.y_axis.title = "Valores"
        
        # Agregar los datos y las categorías al gráfico
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Calcular la posición dos celdas después de la última columna encontrada
        new_column_index = ws.max_column + 2
        new_position = f"{get_column_letter(new_column_index)}1"
        
        # Insertar el gráfico en la hoja de Excel en la nueva posición calculada
        ws.add_chart(chart, new_position)
        
        # Guardar los cambios en el archivo Excel
        wb.save(archivo_excel)
    else:
        print("No se encontraron datos en la hoja 'Grafica' del archivo Excel.")

if __name__ == "__main__":
    # Verificar si se proporcionó la ruta del archivo Excel como argumento
    if len(sys.argv) != 2:
        print("Por favor, proporcione la ruta del archivo Excel como argumento.")
    else:
        archivo_excel = sys.argv[1]
        crear_grafica_desde_excel(archivo_excel)